import os

from openpyxl import load_workbook

# source/ directory lives at the project root, one level above this package
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_DIR = os.path.join(_ROOT, "source")


def list_sources(tenant):
    tenant_dir = os.path.join(SOURCE_DIR, tenant.lower())
    if not os.path.isdir(tenant_dir):
        return []
    results = []
    for root, _, files in os.walk(tenant_dir):
        for f in files:
            if not f.startswith("."):
                results.append(os.path.relpath(os.path.join(root, f), tenant_dir))
    return sorted(results)


def read_excel(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:]]
